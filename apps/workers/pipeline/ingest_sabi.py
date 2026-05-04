"""Carga `docs/sabi_export.xlsx` a la tabla `companies` con tier asignado.

Pipeline:
  1. parse_excel  - lee la hoja "Resultados", valida cabeceras esperadas,
     normaliza valores "n.d." -> None.
  2. validate_rows - rechaza filas rotas (NIF vacio, facturacion negativa,
     valores fuera de tipo) antes de tocar BD. Si hay rotos, levanta
     `IngestValidationError` y NO se escribe nada.
  3. dedup_by_nif - SABI exporta cuentas consolidadas + individuales para
     algunas empresas grandes -> 41 NIFs duplicados en el export inicial.
     Heuristica acordada (Leccion 18): "tier mas alto gana" con empate
     -> primera ocurrencia. La cifra individual de filial pesa mas que la
     consolidada del grupo para un B2B local como DEMIN.
  4. upsert_companies - INSERT ... ON CONFLICT (nif) DO UPDATE. Re-ejecutar
     no duplica. Campos calculados externamente (`ia_fit`, `ia_fit_reason`,
     `research_done_at`, `research_data`) se preservan; los campos de Sabi
     se sobrescriben.
  5. verify_distribution - cuenta filas y distribucion por tier al final.

CLI:
    cd apps/workers
    uv run python -m pipeline.ingest_sabi --env dev
    uv run python -m pipeline.ingest_sabi --env dev --dry-run
    uv run python -m pipeline.ingest_sabi --env prod

Idempotente. Tier T1-T4 segun reglas §8.2 de tasks/todo.md.
"""
from __future__ import annotations

import argparse
import re
import sys
import time
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from sqlalchemy import text

EnvName = Literal["dev", "prod"]

WORKERS_DIR = Path(__file__).resolve().parent.parent
EXCEL_PATH = WORKERS_DIR.parent.parent / "docs" / "sabi_export.xlsx"
SHEET_NAME = "Resultados"

EXPECTED_HEADERS: tuple[str | None, ...] = (
    None,
    "Nombre",
    "Nombre",
    "Código NIF",
    "Localidad",
    "Descripción actividad",
    "Dirección web",
    "Ingresos de explotación\nmil EUR\nÚlt. año disp.",
    "Ingresos de explotación\nmil EUR\nAño - 1",
    "Ingresos de explotación\nmil EUR\nAño - 2",
    "Ingresos de explotación\nmil EUR\nAño - 3",
    "Ingresos de explotación\nmil EUR\nAño - 4",
    "EBITDA\nmil EUR\nÚlt. año disp.",
    "EBITDA\nmil EUR\nAño - 1",
    "EBITDA\nmil EUR\nAño - 2",
    "EBITDA\nmil EUR\nAño - 3",
    "EBITDA\nmil EUR\nAño - 4",
    "Deudas financieras\nmil EUR\nÚlt. año disp.",
    "Tesorería\nmil EUR\nÚlt. año disp.",
)

# Indices de columnas usadas (segun §6.1 schema companies).
COL_NOMBRE = 1
COL_NIF = 3
COL_LOCALIDAD = 4
COL_DESCRIPCION = 5
COL_WEB = 6
COL_REV_Y0 = 7
COL_REV_Y1 = 8

# Patron NIF/CIF espanol relajado: 1 letra + 7-9 alfanumericos. Valida formato
# basico; SABI ya filtro por jurisdiccion espanola, no replicamos validacion
# del digito de control.
NIF_RE = re.compile(r"^[A-Z][A-Z0-9]{7,9}$")

NIL_STRINGS = {"n.d.", "n.a.", "nd", "n/a", ""}

TIER_PRIORITY: dict[str | None, int] = {
    "T1": 4,
    "T2": 3,
    "T3": 2,
    "T4": 1,
    "descartado": 0,
    None: 0,
}


class IngestValidationError(Exception):
    """Lanzada antes de tocar BD si el Excel tiene datos rotos."""


@dataclass(slots=True)
class CompanyRow:
    excel_row: int  # 1-indexed segun openpyxl (fila 1 = header)
    nif: str
    nombre: str
    localidad: str | None
    descripcion: str | None
    web: str | None
    rev_y0_keur: float | None
    rev_y1_keur: float | None
    rev_growth_pct: float | None
    tier: str

    def to_db_dict(self) -> dict[str, Any]:
        return {
            "nif": self.nif,
            "nombre": self.nombre,
            "localidad": self.localidad,
            "descripcion": self.descripcion,
            "web": self.web,
            "rev_y0": self.rev_y0_keur,
            "rev_y1": self.rev_y1_keur,
            "rev_growth": self.rev_growth_pct,
            "tier": self.tier,
        }


def _norm_str(v: Any) -> str | None:
    """Devuelve string limpio o None si vacio o "n.d.". Ignora otros tipos."""
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() in NIL_STRINGS:
        return None
    return s


def _norm_num(v: Any) -> float | None:
    """Convierte a float; "n.d." y similares -> None. Numeros negativos pasan;
    la validacion los rechaza despues si la columna no admite negativos."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lower()
    if s in NIL_STRINGS:
        return None
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return None


def assign_tier(rev_y0_keur: float | None, has_web: bool) -> str:
    """Reglas §8.2 de tasks/todo.md (validadas por humano)."""
    if rev_y0_keur is None or rev_y0_keur < 500 or rev_y0_keur >= 20000:
        return "descartado"
    if has_web:
        if 1000 <= rev_y0_keur < 5000:
            return "T1"
        if 5000 <= rev_y0_keur < 20000:
            return "T2"
        if 500 <= rev_y0_keur < 1000:
            return "T3"
        return "descartado"  # solo si web=True y rev fuera de los rangos T1-T3 (no deberia pasar)
    return "T4"


def _growth_pct(y0: float | None, y1: float | None) -> float | None:
    """Crecimiento year-over-year en %. Solo si ambos numeros y y1 > 0
    (evita division por cero y evita ratios delirantes contra bases ~0)."""
    if y0 is None or y1 is None or y1 <= 0:
        return None
    return round((y0 - y1) / y1 * 100, 2)


def parse_excel(path: Path) -> tuple[list[CompanyRow], list[str]]:
    """Lee el Excel y devuelve (filas validas, errores fatales).

    Errores fatales = problemas que paran la ingesta antes de BD:
      - cabeceras inesperadas
      - NIF vacio/None
      - NIF que no matchea NIF_RE (formato roto)
      - rev_y0 o rev_y1 con valor numerico negativo
    """
    if not path.exists():
        raise FileNotFoundError(f"No existe el Excel SABI: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise IngestValidationError(
            f"Hoja '{SHEET_NAME}' no encontrada en {path.name}. "
            f"Hojas disponibles: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]

    raw = list(ws.iter_rows(values_only=True))
    if not raw:
        raise IngestValidationError(f"Hoja '{SHEET_NAME}' vacia.")

    headers = raw[0]
    errors: list[str] = []

    if len(headers) != len(EXPECTED_HEADERS):
        errors.append(
            f"Numero de columnas inesperado: {len(headers)} vs "
            f"{len(EXPECTED_HEADERS)} esperadas"
        )
    else:
        for i, (got, exp) in enumerate(zip(headers, EXPECTED_HEADERS)):
            if got != exp:
                errors.append(f"Cabecera col {i}: {got!r} != esperado {exp!r}")

    if errors:
        return [], errors

    rows: list[CompanyRow] = []
    seen_nifs: set[str] = set()  # tracking solo para reportar dup, no para excluir aqui
    for excel_idx, raw_row in enumerate(raw[1:], start=2):
        nif_raw = _norm_str(raw_row[COL_NIF])
        nombre = _norm_str(raw_row[COL_NOMBRE])

        if not nif_raw:
            errors.append(f"fila {excel_idx}: NIF vacio")
            continue
        if not NIF_RE.match(nif_raw):
            errors.append(f"fila {excel_idx}: NIF con formato invalido: {nif_raw!r}")
            continue
        if not nombre:
            errors.append(f"fila {excel_idx}: nombre vacio (nif={nif_raw})")
            continue

        rev_y0 = _norm_num(raw_row[COL_REV_Y0])
        rev_y1 = _norm_num(raw_row[COL_REV_Y1])

        if rev_y0 is not None and rev_y0 < 0:
            errors.append(
                f"fila {excel_idx} (nif={nif_raw}): rev_y0 negativo: {rev_y0}"
            )
            continue
        if rev_y1 is not None and rev_y1 < 0:
            errors.append(
                f"fila {excel_idx} (nif={nif_raw}): rev_y1 negativo: {rev_y1}"
            )
            continue

        web = _norm_str(raw_row[COL_WEB])
        descripcion = _norm_str(raw_row[COL_DESCRIPCION])
        localidad = _norm_str(raw_row[COL_LOCALIDAD])

        tier = assign_tier(rev_y0, has_web=web is not None)

        rows.append(
            CompanyRow(
                excel_row=excel_idx,
                nif=nif_raw,
                nombre=nombre,
                localidad=localidad,
                descripcion=descripcion,
                web=web,
                rev_y0_keur=rev_y0,
                rev_y1_keur=rev_y1,
                rev_growth_pct=_growth_pct(rev_y0, rev_y1),
                tier=tier,
            )
        )
        seen_nifs.add(nif_raw)

    return rows, errors


def dedup_by_nif(rows: list[CompanyRow]) -> tuple[list[CompanyRow], list[tuple[str, str, str]]]:
    """Aplica heuristica "tier mas alto gana, empate -> primera ocurrencia".

    Devuelve (filas deduplicadas, reporte de decisiones tomadas) donde el
    reporte es lista de tuplas `(nif, tier_kept, tier_dropped)` solo para
    los NIFs que tuvieron duplicado.
    """
    by_nif: dict[str, CompanyRow] = {}
    decisions: list[tuple[str, str, str]] = []
    for row in rows:
        existing = by_nif.get(row.nif)
        if existing is None:
            by_nif[row.nif] = row
            continue
        # Hay duplicado. Comparar prioridad de tier.
        prio_new = TIER_PRIORITY[row.tier]
        prio_old = TIER_PRIORITY[existing.tier]
        if prio_new > prio_old:
            decisions.append((row.nif, row.tier, existing.tier))
            by_nif[row.nif] = row
        else:
            # Empate o nuevo peor: nos quedamos con el existente (primera ocurrencia).
            decisions.append((row.nif, existing.tier, row.tier))
    return list(by_nif.values()), decisions


_UPSERT_SQL = text(
    """
    insert into companies (
        nif, nombre, localidad, descripcion, web,
        rev_y0_keur, rev_y1_keur, rev_growth_pct, tier
    ) values (
        :nif, :nombre, :localidad, :descripcion, :web,
        :rev_y0, :rev_y1, :rev_growth, :tier
    )
    on conflict (nif) do update set
        nombre         = excluded.nombre,
        localidad      = excluded.localidad,
        descripcion    = excluded.descripcion,
        web            = excluded.web,
        rev_y0_keur    = excluded.rev_y0_keur,
        rev_y1_keur    = excluded.rev_y1_keur,
        rev_growth_pct = excluded.rev_growth_pct,
        tier           = excluded.tier
    -- ia_fit, ia_fit_reason, research_done_at, research_data, created_at
    -- NO se tocan: los rellenan workers posteriores.
    """
)


def upsert_companies(env: EnvName, rows: Iterable[CompanyRow], batch_size: int = 500) -> int:
    """Aplica UPSERT en lotes. Devuelve nº de filas procesadas."""
    # Import lazy para no forzar carga de settings al importar el modulo.
    from shared.db import get_session  # noqa: PLC0415

    payload = [r.to_db_dict() for r in rows]
    n = 0
    with get_session(env) as s:
        for i in range(0, len(payload), batch_size):
            batch = payload[i : i + batch_size]
            s.execute(_UPSERT_SQL, batch)
            n += len(batch)
    return n


def verify_distribution(env: EnvName) -> dict[str, int]:
    """Cuenta companies por tier en BD. Util para smoke y reporte."""
    from shared.db import get_session  # noqa: PLC0415

    with get_session(env) as s:
        rs = s.execute(
            text(
                "select coalesce(tier, 'NULL') as tier, count(*) as n "
                "from companies group by tier"
            )
        ).all()
    return {row[0]: int(row[1]) for row in rs}


def _print_distribution(label: str, dist: dict[str, int]) -> None:
    total = sum(dist.values())
    print(f"  {label}: total={total}")
    for tier in ("T1", "T2", "T3", "T4", "descartado"):
        n = dist.get(tier, 0)
        pct = (n / total * 100) if total else 0
        print(f"    {tier:<11} {n:>5}  ({pct:5.1f}%)")
    extras = {k: v for k, v in dist.items() if k not in {"T1", "T2", "T3", "T4", "descartado"}}
    for k, v in extras.items():
        print(f"    {k:<11} {v:>5}  (otros)")


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Ingesta SABI -> companies")
    p.add_argument("--env", choices=("dev", "prod"), required=True)
    p.add_argument("--dry-run", action="store_true",
                   help="Parse + valida + dedup, NO escribe BD.")
    p.add_argument("--excel", type=Path, default=EXCEL_PATH,
                   help=f"Path del Excel (default {EXCEL_PATH}).")
    args = p.parse_args(argv)
    env: EnvName = args.env  # type: ignore[assignment]

    print("=" * 76)
    print(f"ingest_sabi  env={env}  dry_run={args.dry_run}  excel={args.excel.name}")
    print("=" * 76)

    t0 = time.monotonic()
    rows, errors = parse_excel(args.excel)
    if errors:
        print(f"PARADA: {len(errors)} errores fatales detectados antes de tocar BD")
        for e in errors[:25]:
            print(f"  - {e}")
        if len(errors) > 25:
            print(f"  ... ({len(errors) - 25} mas)")
        return 1

    print(f"[parse] {len(rows)} filas validas en {time.monotonic() - t0:.1f}s")

    deduped, decisions = dedup_by_nif(rows)
    n_dups = len(rows) - len(deduped)
    print(f"[dedup] {len(deduped)} NIFs unicos; eliminados {n_dups} duplicados")
    if decisions:
        # Conteo de transitions
        kept = {}
        for _, k, _d in decisions:
            kept[k] = kept.get(k, 0) + 1
        print(f"  decisiones tomadas: {len(decisions)} duplicados resueltos")
        print(f"  tier finalmente conservado en cada caso: {kept}")

    # Distribucion in-memory antes de tocar BD.
    dist_mem = {}
    for r in deduped:
        dist_mem[r.tier] = dist_mem.get(r.tier, 0) + 1
    print()
    _print_distribution("distribucion calculada (pre-UPSERT)", dist_mem)

    if args.dry_run:
        print()
        print("DRY-RUN: no se escribe BD.")
        return 0

    print()
    print(f"[upsert] aplicando UPSERT a {env}...")
    t1 = time.monotonic()
    n = upsert_companies(env, deduped)
    print(f"[upsert] {n} filas procesadas en {time.monotonic() - t1:.1f}s")

    print()
    dist_db = verify_distribution(env)
    _print_distribution(f"distribucion en BD ({env})", dist_db)

    if dist_db.get("T1", 0) + dist_db.get("T2", 0) + dist_db.get("T3", 0) + dist_db.get("T4", 0) == 0:
        print("ALERTA: 0 filas accionables en BD. Revisar.")
        return 2
    print()
    print(f"OK: ingest_sabi completado en {time.monotonic() - t0:.1f}s")
    return 0


if __name__ == "__main__":
    sys.exit(main())
